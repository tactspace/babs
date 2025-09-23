from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from typing import Dict, List, Tuple, Any
import os
import json
from models import RouteRequest, RouteResult, Driver, SingleRouteRequest, SingleRouteResponse, SingleRouteWithSegments, MultiRouteWithSegments
from trucks import load_truck_specs
from charging_stations import load_charging_stations
from route_calculator import calculate_detailed_route, calculate_multi_route
from optimizer import optimize_routes
from optily import plan_route
from optily_multiroute import calculate_multi_route_with_swaps

app = FastAPI(title="E-Truck Routing Optimizer")

# Enable CORS for local frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load data at startup
truck_specs = {}
charging_stations = []
drivers: dict[str, Driver] = {}

@app.on_event("startup")
async def startup_event():
    global truck_specs, charging_stations, drivers
    
    # Load truck specifications
    truck_specs = load_truck_specs("data/truck_specs.csv")
    
    # Load charging stations
    charging_stations = load_charging_stations("data/public_charge_points.csv")
    
    # Load drivers (mock + from xlsx if available)
    try:
        from openpyxl import load_workbook
        wb = load_workbook("data/drivers_distribution.xlsx")
        ws = wb.active
        # Expect headers in first row: id, name
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not row:
                continue
            did = str(row[0]) if row[0] is not None else f"drv_{i}"
            name = str(row[1]) if len(row) > 1 and row[1] is not None else f"Driver {i}"
            drivers[did] = Driver(id=did, name=name)
    except Exception:
        # Fallback mock drivers
        
        print("Error loading drivers")


@app.get("/")
async def root():
    return {"message": "E-Truck Routing Optimizer API"}


@app.get("/charging-stations")
async def get_charging_stations(
    country: str = None,
    truck_suitable_only: bool = False,
    limit: int = 100
) -> List:
    """Get charging stations with optional filters"""
    filtered = charging_stations
    
    if country:
        filtered = [s for s in filtered if s.country == country]
    
    if truck_suitable_only:
        filtered = [s for s in filtered if s.truck_suitability == "yes"]
    
    # Return limited number of stations
    return [station.dict() for station in filtered[:limit]]

@app.post("/get-optimal-route", response_model=SingleRouteResponse)
async def get_optimal_route(request: SingleRouteRequest):
    """Get a simple route between two points using TomTom API"""
    try:
        start_point = (request.start_lat, request.start_lng)
        end_point = (request.end_lat, request.end_lng)
        route_name = request.route_name
        
        # Call TomTom API
        from tomtom import get_route
        route_data = get_route(start_point, end_point)

        # export to json
        with open(f"{route_name}.json", "w") as f:
            json.dump(route_data, f)
        
        if not route_data:
            return SingleRouteResponse(
                distance_km=0,
                route_name=route_name,
                duration_minutes=0,
                coordinates=[],
                success=False,
                message="Something went wrong. Could not calculate route"
            )
        
        # Convert coordinates to the format expected by frontend
        coordinates = []
        for point in route_data["coordinates"]:
            coordinates.append({
                "lat": point["latitude"],
                "lng": point["longitude"]
            })
        
        return SingleRouteResponse(
            distance_km=route_data["distance"] / 1000,  # Convert meters to km
            duration_minutes=route_data["duration"] / 60,  # Convert seconds to minutes
            coordinates=coordinates,
            success=True,
            message="Route calculated successfully",
            route_name=route_name
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




# Refactoring needed below
@app.get("/trucks")
async def get_trucks() -> Dict:
    """Get available truck models"""
    return {model: truck.dict() for model, truck in truck_specs.items()}
@app.get("/drivers")
async def get_drivers() -> Dict:
    return {k: v.dict() for k, v in drivers.items()}


@app.get("/charging-stations/{station_id}")
async def get_charging_station(station_id: int):
    """Get details of a specific charging station"""
    for station in charging_stations:
        if station.id == station_id:
            return station.dict()
    
    raise HTTPException(status_code=404, detail="Charging station not found")


@app.post("/route", response_model=RouteResult)
async def calculate_route(request: RouteRequest):
    """Calculate optimal route for an e-truck"""
    # Validate truck model
    if request.truck_model not in truck_specs:
        raise HTTPException(
            status_code=400, 
            detail=f"Unknown truck model: {request.truck_model}. Available models: {list(truck_specs.keys())}"
        )
    
    # Find optimal route
    try:
        result = find_optimal_route(request)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/optimize", response_model=RouteResult)
async def optimize_route(request: RouteRequest):
    """Optimize route with driver swaps and charging stations"""
    # Validate truck model
    if request.truck_model not in truck_specs:
        raise HTTPException(
            status_code=400, 
            detail=f"Unknown truck model: {request.truck_model}. Available models: {list(truck_specs.keys())}"
        )
    
    # Convert request to format expected by optimizer
    routes = [{
        "start_coord": {"latitude": request.start_point[0], "longitude": request.start_point[1]},
        "end_coord": {"latitude": request.end_point[0], "longitude": request.end_point[1]}
    }]
    
    # Create driver instances based on request
    num_drivers = request.num_drivers or 1
    optimizer_drivers = []
    for i in range(num_drivers):
        driver_id = request.driver_ids[i] if request.driver_ids and i < len(request.driver_ids) else i+1
        optimizer_drivers.append(Driver(
            id=driver_id, 
            name=f"Driver {driver_id}",
            home_location=request.start_point
        ))
    
    try:
        # Run optimization
        result = optimize_routes(routes, charging_stations, optimizer_drivers)
        
        # Convert result to RouteResult format
        route_result = RouteResult(
            total_distance=result["total_distance"] * 1000,  # convert km to meters
            total_duration=sum(iter["time_elapsed_minutes"] * 60 for iter in result["iterations"]),
            driving_duration=sum(iter.get("driving_time_seconds", 0) for iter in result["iterations"]),
            total_energy_consumption=sum(iter.get("energy_consumption", 0) for iter in result["iterations"]),
            total_cost=sum(iter["sum_cost"] for iter in result["iterations"]),
            route_segments=[],  # Would need to convert iterations to route segments
            driver_breaks=[],  # Would need to extract from iterations
            charging_stops=[],  # Would need to extract from iterations
            driver_swaps=[],  # Would need to convert from result["truck_swaps"]
            feasible=True
        )
        
        return route_result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/calculate-costs", response_model=SingleRouteWithSegments)
async def calculate_costs(request: SingleRouteRequest, truck_model: str = None, starting_battery_kwh: float = None):
    """Calculate detailed route costs with segments and charging stops"""
    try:
        # Validate truck model if provided
        if truck_model and truck_model not in truck_specs:
            raise HTTPException(
                status_code=400, 
                detail=f"Unknown truck model: {truck_model}. Available models: {list(truck_specs.keys())}"
            )
        
        # Call the enhanced route planner
        result = plan_route(request, truck_model, starting_battery_kwh)
        
        if not result.success:
            raise HTTPException(status_code=400, detail=result.message)
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating route costs: {str(e)}")


@app.post("/compare-costs", response_model=MultiRouteWithSegments)
async def compare_costs(route_requests: List[SingleRouteRequest], starting_battery_kwh: float = None):
    """
    Compare costs for multiple routes with and without driver swapping optimization.
    
    This endpoint:
    1. Calculates individual routes using the existing plan_route function
    2. Runs optimization to find beneficial driver swaps
    3. Returns detailed comparison between base and optimized costs
    
    Args:
        route_requests: List of SingleRouteRequest objects
        starting_battery_kwh: Starting battery level (optional)
        
    Returns:
        MultiRouteWithSegments with detailed cost comparisons and optimization results
    """
    try:
        # Validate input
        if not route_requests:
            raise HTTPException(status_code=400, detail="At least one route request is required")
        
        if len(route_requests) < 2:
            raise HTTPException(status_code=400, detail="At least two routes are required for comparison")
        
        # Call the multi-route optimization function
        result = calculate_multi_route_with_swaps(route_requests, starting_battery_kwh)
        with open(f"simple_routes.json", "w") as f:
            f.write(f"Result:\n {result}")
        
        if not result.success:
            raise HTTPException(status_code=400, detail=result.message)
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error comparing route costs: {str(e)}")


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "loaded_trucks": len(truck_specs),
        "loaded_charging_stations": len(charging_stations)
    }


# Add new models for the detailed route API
from pydantic import BaseModel
from typing import List, Optional, Tuple

class DetailedRouteRequest(BaseModel):
    start_point: List[float]
    end_point: List[float]
    truck_type: Optional[str] = "electric"

class MultiRouteRequest(BaseModel):
    routes: List[DetailedRouteRequest]

@app.post("/detailed-route")
async def get_detailed_route(request: DetailedRouteRequest):
    """Calculate a route with detailed cost breakdown"""
    try:
        start_point = (request.start_point[0], request.start_point[1])
        end_point = (request.end_point[0], request.end_point[1])
        
        result = calculate_detailed_route(
            start_point=start_point,
            end_point=end_point,
            truck_type=request.truck_type
        )
        
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/multi-route")
async def calculate_multiple_routes(request: MultiRouteRequest):
    """Calculate multiple routes with detailed cost information and comparison with optimized version"""
    try:
        # Step 1: Calculate base routes using route_calculator
        base_routes = []
        for route in request.routes:
            base_routes.append({
                "start_point": route.start_point,
                "end_point": route.end_point,
                "truck_type": route.truck_type
            })
        
        base_result = calculate_multi_route(base_routes)
        
        # Step 2: Calculate optimized routes for each route separately
        for i, (route, base_route_result) in enumerate(zip(request.routes, base_result["routes"])):
            # Create a single route for optimization
            opt_route = [{
                "start_coord": {"latitude": route.start_point[0], "longitude": route.start_point[1]},
                "end_coord": {"latitude": route.end_point[0], "longitude": route.end_point[1]}
            }]
            
            # Create 2 drivers for optimization
            opt_drivers = [
                Driver(
                    id=1,
                    name="Driver 1",
                    home_location=(route.start_point[0], route.start_point[1])
                ),
                Driver(
                    id=2,
                    name="Driver 2",
                    home_location=(route.start_point[0], route.start_point[1])
                )
            ]
            
            try:
                # Run optimization for this single route
                opt_result = optimize_routes(opt_route, charging_stations, opt_drivers)
                
                # Calculate route-specific comparison
                route_comparison = {
                    "base": {
                        "total_cost": base_route_result["total_cost"],
                        "total_duration": base_route_result["total_duration"],
                        "total_energy": base_route_result["total_energy_consumption"],
                        "total_distance": base_route_result["total_distance"]
                    },
                    "optimized": {
                        "total_cost": sum(iter.get("sum_cost", 0) for iter in opt_result.get("iterations", [])),
                        "total_duration": sum(iter.get("time_elapsed_minutes", 0) * 60 for iter in opt_result.get("iterations", [])),
                        "total_energy": opt_result.get("total_distance", 0) * 1.2,  # Estimate energy based on distance
                        "total_distance": opt_result.get("total_distance", 0) * 1000  # convert km to meters
                    }
                }
                
                # Calculate savings
                route_comparison["savings"] = {
                    "cost": route_comparison["base"]["total_cost"] - route_comparison["optimized"]["total_cost"],
                    "cost_percentage": ((route_comparison["base"]["total_cost"] - route_comparison["optimized"]["total_cost"]) / route_comparison["base"]["total_cost"]) * 100 if route_comparison["base"]["total_cost"] > 0 else 0,
                    "duration": route_comparison["base"]["total_duration"] - route_comparison["optimized"]["total_duration"],
                    "duration_percentage": ((route_comparison["base"]["total_duration"] - route_comparison["optimized"]["total_duration"]) / route_comparison["base"]["total_duration"]) * 100 if route_comparison["base"]["total_duration"] > 0 else 0,
                    "energy": route_comparison["base"]["total_energy"] - route_comparison["optimized"]["total_energy"],
                    "energy_percentage": ((route_comparison["base"]["total_energy"] - route_comparison["optimized"]["total_energy"]) / route_comparison["base"]["total_energy"]) * 100 if route_comparison["base"]["total_energy"] > 0 else 0
                }
                
                # Add comparison to the route result
                base_route_result["comparison"] = route_comparison
                
            except Exception as e:
                print(f"Error optimizing route {i}: {e}")
                # Continue with other routes if one fails
        
        # Step 3: Create overall comparison
        overall_comparison = {
            "base": {
                "total_cost": base_result["total_cost"],
                "total_duration": base_result["total_duration"],
                "total_energy": sum(r["total_energy_consumption"] for r in base_result["routes"]),
                "total_distance": base_result["total_distance"]
            },
            "optimized": {
                "total_cost": sum(r.get("comparison", {}).get("optimized", {}).get("total_cost", 0) for r in base_result["routes"]),
                "total_duration": sum(r.get("comparison", {}).get("optimized", {}).get("total_duration", 0) for r in base_result["routes"]),
                "total_energy": sum(r.get("comparison", {}).get("optimized", {}).get("total_energy", 0) for r in base_result["routes"]),
                "total_distance": sum(r.get("comparison", {}).get("optimized", {}).get("total_distance", 0) for r in base_result["routes"])
            }
        }
        
        # Calculate overall savings
        overall_comparison["savings"] = {
            "cost": overall_comparison["base"]["total_cost"] - overall_comparison["optimized"]["total_cost"],
            "cost_percentage": ((overall_comparison["base"]["total_cost"] - overall_comparison["optimized"]["total_cost"]) / overall_comparison["base"]["total_cost"]) * 100 if overall_comparison["base"]["total_cost"] > 0 else 0,
            "duration": overall_comparison["base"]["total_duration"] - overall_comparison["optimized"]["total_duration"],
            "duration_percentage": ((overall_comparison["base"]["total_duration"] - overall_comparison["optimized"]["total_duration"]) / overall_comparison["base"]["total_duration"]) * 100 if overall_comparison["base"]["total_duration"] > 0 else 0,
            "energy": overall_comparison["base"]["total_energy"] - overall_comparison["optimized"]["total_energy"],
            "energy_percentage": ((overall_comparison["base"]["total_energy"] - overall_comparison["optimized"]["total_energy"]) / overall_comparison["base"]["total_energy"]) * 100 if overall_comparison["base"]["total_energy"] > 0 else 0
        }
        
        # Add overall comparison to result
        base_result["comparison"] = overall_comparison
        
        return base_result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)